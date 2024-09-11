from openpyxl import load_workbook

archivo_excel = "../Include/asistencia.xlsx"  # Ruta del archivo de Excel
libro = load_workbook(archivo_excel)
hoja = libro["Asistencia"]  # Obtener la hoja "Asistencia"

# Comenzar en la fila 2, suponiendo que la fila 1 contiene encabezados
nueva_fila = 2
while hoja.cell(row=nueva_fila, column=1).value is not None:
    nueva_fila += 1  # Buscar la primera fila vacía después de los encabezados

# Solicitar datos al usuario
nombre = input("Ingresa tu nombre: ")
fecha = input("Ingresa la fecha (yyyy-mm-dd): ")
hora = input("Ingresa la hora de entrada (hh:mm): ")

# Escribir los datos en la primera fila vacía después de los encabezados
hoja.cell(row=nueva_fila, column=1).value = nombre
hoja.cell(row=nueva_fila, column=2).value = fecha
hoja.cell(row=nueva_fila, column=3).value = hora

# Guardar los cambios en el archivo Excel
libro.save(archivo_excel)

print(f"Datos guardados exitosamente en la fila {nueva_fila}.")
